import io
import logging
import re
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def import_from_workbook(
    file_obj,
    Consignment=None,
    db=None,
    normalize_consignment_number=None,
    normalize_status=None,
    normalize_indian_pincode=None,
):
    """Import consignments from an uploaded workbook file-like object.

    Returns (added_count, skipped_count).
    Accepts optional `Consignment` and `db` to allow caller to pass patched objects for testing.
    """
    if Consignment is None:
        from app.models import Consignment as Consignment
    if db is None:
        from app.models import db as db

    # use repository helpers where possible
    from app.services import consignment_repo as repo

    if (
        normalize_consignment_number is None
        or normalize_status is None
        or normalize_indian_pincode is None
    ):
        from app.services.logistics import (
            normalize_consignment_number as _ncn,
            normalize_status as _ns,
            normalize_indian_pincode as _nip,
        )

        normalize_consignment_number = normalize_consignment_number or _ncn
        normalize_status = normalize_status or _ns
        normalize_indian_pincode = normalize_indian_pincode or _nip

    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        raise ValueError("Excel file is empty.")

    normalized_headers = [_normalize_header(cell) for cell in header_cells]
    header_index = {name: idx for idx, name in enumerate(normalized_headers) if name}

    consignment_idx = header_index.get("consignment_number")
    status_idx = header_index.get("status")
    pickup_address_idx = header_index.get("pickup_address")
    pickup_pincode_idx = header_index.get("pickup_pincode")
    pickup_tag_idx = header_index.get("pickup_tag")
    pickup_date_idx = header_index.get("pickup_date")
    drop_address_idx = header_index.get("drop_address")
    drop_pincode_idx = header_index.get("drop_pincode")
    drop_tag_idx = header_index.get("drop_tag")
    drop_date_idx = header_index.get("drop_date")
    eta_idx = header_index.get("eta")

    if None in (consignment_idx, status_idx):
        raise ValueError("Required headers: consignment_number, status")

    # Prefer the injected Consignment (useful for tests that patch the model)
    if Consignment is not None:
        try:
            existing_numbers = {
                c[0]
                for c in Consignment.query.with_entities(
                    Consignment.consignment_number
                ).all()
            }
        except Exception:
            # If the injected Consignment cannot be queried (tests may not patch repo),
            # fall back to repository helper which uses the real model.
            existing_numbers = repo.query_existing_numbers()
    else:
        existing_numbers = repo.query_existing_numbers()
    file_seen = set()
    added_count = 0
    skipped_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        consignment_number = normalize_consignment_number(row[consignment_idx])
        status = normalize_status(row[status_idx])
        pickup_address = str(
            row[pickup_address_idx]
            if pickup_address_idx is not None and row[pickup_address_idx] is not None
            else ""
        ).strip()
        pickup_pincode = normalize_indian_pincode(
            row[pickup_pincode_idx]
            if pickup_pincode_idx is not None and row[pickup_pincode_idx] is not None
            else "",
            "pickup_pincode",
        )
        pickup_tag = str(
            row[pickup_tag_idx]
            if pickup_tag_idx is not None and row[pickup_tag_idx] is not None
            else ""
        ).strip()
        pickup_date = str(
            row[pickup_date_idx]
            if pickup_date_idx is not None and row[pickup_date_idx] is not None
            else ""
        ).strip()
        drop_address = str(
            row[drop_address_idx]
            if drop_address_idx is not None and row[drop_address_idx] is not None
            else ""
        ).strip()
        drop_pincode = normalize_indian_pincode(
            row[drop_pincode_idx]
            if drop_pincode_idx is not None and row[drop_pincode_idx] is not None
            else "",
            "drop_pincode",
        )
        drop_tag = str(
            row[drop_tag_idx]
            if drop_tag_idx is not None and row[drop_tag_idx] is not None
            else ""
        ).strip()
        drop_date = str(
            row[drop_date_idx]
            if drop_date_idx is not None and row[drop_date_idx] is not None
            else ""
        ).strip()
        eta = str(
            row[eta_idx] if eta_idx is not None and row[eta_idx] is not None else ""
        ).strip()

        if consignment_number in existing_numbers or consignment_number in file_seen:
            skipped_count += 1
            continue

        consignment = Consignment(
            consignment_number=consignment_number,
            status=status,
            pickup_address=pickup_address,
            pickup_pincode=pickup_pincode,
            pickup_tag=pickup_tag,
            pickup_date=pickup_date,
            drop_address=drop_address,
            drop_pincode=drop_pincode,
            drop_tag=drop_tag,
            drop_date=drop_date,
            eta=eta,
        )

        # Prefer the caller-provided `db` for session operations (tests may inject a fake db).
        if db is not None:
            db.session.add(consignment)
        else:
            repo.add(consignment)
        file_seen.add(consignment_number)
        existing_numbers.add(consignment_number)
        added_count += 1

    # Commit all added rows within a single transaction boundary.
    from app.db.session import transaction

    with transaction(db):
        # all adds were performed above (either via injected `db.session.add` or repo.add)
        pass
    return added_count, skipped_count


def generate_import_template_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import Template"

    sheet.append(
        [
            "consignment_number",
            "status",
            "pickup_address",
            "pickup_pincode",
            "pickup_tag",
            "pickup_date",
            "drop_address",
            "drop_pincode",
            "drop_tag",
            "drop_date",
        ]
    )

    sheet.append(
        [
            "CN001",
            "In Transit",
            "123 Main Street, New Delhi",
            "110017",
            "PICKUP-001",
            "2026-05-10",
            "456 Marine Drive, Mumbai",
            "400001",
            "DROP-001",
            "2026-05-12",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_rows_to_workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Internal Consignments"

    sheet.append(
        [
            "consignment_number",
            "status",
            "pickup_tag",
            "drop_pincode",
            "pickup_date",
            "drop_date",
            "pickup_address",
            "drop_address",
        ]
    )

    for row in rows:
        sheet.append(
            [
                row.consignment_number,
                row.status,
                getattr(row, "pickup_tag", ""),
                row.drop_pincode,
                getattr(row, "pickup_date", ""),
                getattr(row, "drop_date", ""),
                getattr(row, "pickup_address", ""),
                getattr(row, "drop_address", ""),
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
